"""Parse Mailchimp audience exports (CSV or Excel) and import into mailing lists."""

from __future__ import annotations

import csv
import io
import secrets
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import HTTPException, status
from sqlalchemy.orm import Session

from app.models.models import MailingList, MailingSubscriber

_MAILCHIMP_EMAIL_KEYS = ("email address", "email", "email_address")
_MAILCHIMP_FIRST_NAME_KEYS = ("first name", "fname", "first_name")
_MAILCHIMP_LAST_NAME_KEYS = ("last name", "lname", "last_name")
_MAILCHIMP_STATUS_KEYS = ("status", "email marketing status")
_MAILCHIMP_CONSENT_KEYS = ("timestamp signup", "optin time", "date added")

_ALLOWED_EXTENSIONS = (".csv", ".xlsx", ".xlsm")


def normalize_mailchimp_status(value: str | None) -> str:
    raw = (value or "").strip().lower()
    if raw in {"", "subscribed"}:
        return "subscribed"
    if raw in {"unsubscribed", "archived"}:
        return "unsubscribed"
    if raw in {"cleaned", "non-subscribed", "nonsubscribed", "pending", "transactional"}:
        return "cleaned"
    return "subscribed"


def parse_datetime(value: str | None) -> datetime | None:
    text = (value or "").strip()
    if not text:
        return None
    for pattern in (
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%S",
    ):
        try:
            parsed = datetime.strptime(text, pattern)
            if parsed.tzinfo is None:
                return parsed.replace(tzinfo=timezone.utc)
            return parsed.astimezone(timezone.utc)
        except ValueError:
            continue
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            return parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc)
    except ValueError:
        return None


def cell(row: dict[str, str], *keys: str) -> str:
    normalized = {str(k).strip().lower(): (v or "") for k, v in row.items()}
    for key in keys:
        if key in normalized and normalized[key].strip():
            return normalized[key].strip()
    return ""


def build_full_name(row: dict[str, str]) -> str | None:
    first_name = cell(row, *_MAILCHIMP_FIRST_NAME_KEYS)
    last_name = cell(row, *_MAILCHIMP_LAST_NAME_KEYS)
    full_name = " ".join(part for part in [first_name, last_name] if part).strip()
    return full_name or None


def _format_excel_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _decode_csv_bytes(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Could not decode CSV file")


def rows_from_csv(data: bytes) -> list[dict[str, str]]:
    csv_text = _decode_csv_bytes(data)
    reader = csv.DictReader(io.StringIO(csv_text))
    if not reader.fieldnames:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="CSV header row is missing")
    return [dict(row) for row in reader]


def rows_from_xlsx(data: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel import is not available on the server (openpyxl missing).",
        ) from exc

    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Could not read Excel file: {exc}",
        ) from exc

    try:
        sheet = workbook.active
        row_iter = sheet.iter_rows(values_only=True)
        try:
            header_cells = next(row_iter)
        except StopIteration:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel file is empty")

        headers = [_format_excel_value(h) for h in header_cells]
        if not any(headers):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel header row is missing")

        rows: list[dict[str, str]] = []
        for cells in row_iter:
            if cells is None or all(
                c is None or (isinstance(c, str) and not c.strip()) for c in cells
            ):
                continue
            record: dict[str, str] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = cells[index] if index < len(cells) else None
                record[header] = _format_excel_value(value)
            rows.append(record)
        return rows
    finally:
        workbook.close()


def parse_mailchimp_audience_file(filename: str | None, data: bytes) -> list[dict[str, str]]:
    if not data:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File is empty")

    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        return rows_from_csv(data)
    if lower.endswith((".xlsx", ".xlsm")):
        return rows_from_xlsx(data)
    if lower.endswith(".xls"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Legacy .xls is not supported. In Excel or Mailchimp, save/export as .xlsx or .csv.",
        )
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail=f"Upload a Mailchimp export ({', '.join(_ALLOWED_EXTENSIONS)})",
    )


def ensure_target_list(
    db: Session,
    *,
    existing_list_id: int | None,
    list_name: str | None,
    filename: str | None,
) -> MailingList:
    if existing_list_id is not None:
        mailing_list = db.get(MailingList, existing_list_id)
        if not mailing_list:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audience not found")
        return mailing_list

    resolved_name = (list_name or "").strip()
    if not resolved_name:
        stem = Path(filename or "mailchimp-import").stem.replace("_", " ").replace("-", " ").strip()
        resolved_name = stem or "Mailchimp Import"

    mailing_list = MailingList(name=resolved_name, description="Imported from Mailchimp")
    db.add(mailing_list)
    db.commit()
    db.refresh(mailing_list)
    return mailing_list


def import_mailchimp_rows(
    db: Session,
    mailing_list: MailingList,
    rows: list[dict[str, str]],
    *,
    source_label: str,
) -> tuple[int, int, int]:
    """Returns (created, updated, skipped)."""
    created = 0
    updated = 0
    skipped = 0

    for row in rows:
        email = cell(row, *_MAILCHIMP_EMAIL_KEYS).lower()
        if not email:
            skipped += 1
            continue

        full_name = build_full_name(row)
        status_value = normalize_mailchimp_status(cell(row, *_MAILCHIMP_STATUS_KEYS))
        consent_at = parse_datetime(cell(row, *_MAILCHIMP_CONSENT_KEYS))
        subscriber = (
            db.query(MailingSubscriber)
            .filter(MailingSubscriber.list_id == mailing_list.id, MailingSubscriber.email == email)
            .first()
        )

        if subscriber:
            subscriber.full_name = full_name or subscriber.full_name
            subscriber.status = status_value
            subscriber.consent_source = subscriber.consent_source or source_label
            subscriber.consent_at = subscriber.consent_at or consent_at or datetime.now(timezone.utc)
            subscriber.unsubscribed_at = datetime.now(timezone.utc) if status_value == "unsubscribed" else None
            updated += 1
            continue

        db.add(
            MailingSubscriber(
                list_id=mailing_list.id,
                email=email,
                full_name=full_name,
                status=status_value,
                consent_source=source_label,
                consent_at=consent_at or datetime.now(timezone.utc),
                unsubscribed_at=datetime.now(timezone.utc) if status_value == "unsubscribed" else None,
                unsubscribe_token=secrets.token_urlsafe(24),
            )
        )
        created += 1

    db.commit()
    return created, updated, skipped
